# See this guide on how to implement these action:
# https://rasa.com/docs/rasa/custom-actions

from typing import Any, Coroutine, Text, Dict, List
from rasa_sdk import Action, Tracker, FormValidationAction
from rasa_sdk.executor import CollectingDispatcher
from rasa_sdk.events import SlotSet, AllSlotsReset,UserUtteranceReverted
from rasa_sdk.types import DomainDict
import openpyxl
import re
import random
from datetime import datetime, timedelta
import mysql.connector as mc

orders = []
# *support if i type piiiizzzzzaaaa or pizzs so does  it work?*

class ActionSayData(Action):
    FOOD_PRICES = {
        "dosa": 80,
        "masala-dosa": 100,
        "plain-dosa": 80,
        "masala_dosa": 100,
        "plain_dosa": 80,
        "idli": 15,
        "vada": 20,
        "utappam": 60,
        "uttapam":60,
        "pizza": 110,
        "burger": 70,
        "pasta": 70,
        "sandwich": 80,
        "sandwitch":80,
        "Franky":60,
        "fries":40
    }

    def name(self) -> Text:
        return "action_say_data"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: DomainDict) -> Coroutine[Any, Any, List[Dict[Text, Any]]]:
        food = tracker.get_slot('Food')
        quantity = tracker.get_slot('Quantity')
        food_lower = food.lower()  # Convert input to lowercase for case-insensitivity

        order = {"food": food_lower, "quantity": quantity}
        orders.append(order)

        latest_message = tracker.latest_message
        text = latest_message.get('text', '')

        if text.lower() == 'yes' or text.lower() == 'y':
            return [SlotSet("Food", None), SlotSet("Quantity", None), SlotSet("requested_slot")]
        else:
            name = tracker.get_slot('Name')
            phone_number = tracker.get_slot('PhoneNumber')[0]
            address = ','.join(tracker.get_slot('Address'))
            order_generation_time = datetime.now()
            order_id = str(random.randint(1000, 9999))
            print(order_generation_time)
            price = sum(self.FOOD_PRICES[item['food']] * int(item['quantity']) for item in orders)

            order_text = ', '.join([f"{item['quantity']} {item['food']}" for item in orders])

            dispatcher.utter_message(f"Hey {name}, You have ordered {order_text} and your contact number is {phone_number} and your bill is {price}/- only")
            dispatcher.utter_message(f"Thank you for visiting. Your OrderId is {order_id} and your order will be ready in 15 minutes")

            print(order_text)
            print(price)
            connection = mc.connect(
            host="localhost",
            user="Rasa_database",
            password="Kutub@123",
            database="rasa"
            )
            cursor = connection.cursor()

            # Insert data into the database
            insert_query = "INSERT INTO information (orderId,name, phone_number, address,ORDERS,price,TimeStamp) VALUES (%s, %s, %s,%s, %s, %s,%s)"
            data = (order_id,name, phone_number, address,order_text,price,order_generation_time)

            cursor.execute(insert_query, data)
            connection.commit()

            # Close the connection
            cursor.close()
            connection.close()

            # Create or open the Excel file
            try:
                workbook = openpyxl.load_workbook("food_order.xlsx")  # Try loading existing file
            except FileNotFoundError:
                workbook = openpyxl.Workbook()  # Create a new file if not found

            sheet = workbook.active  # Get the active worksheet
            next_row = sheet.max_row + 1  # Find the next empty row
            sheet.cell(row=next_row, column=1).value = order_id
            sheet.cell(row=next_row, column=2).value = name
            sheet.cell(row=next_row, column=3).value = phone_number
            sheet.cell(row=next_row, column=4).value = order_text
            sheet.cell(row=next_row, column=5).value = address
            sheet.cell(row=next_row, column=6).value = order_generation_time

            workbook.save("food_order.xlsx")  # Save the updated Excel file
            orders.clear()  # Clear the list of orders

            return [AllSlotsReset(), SlotSet("requested_slot", None)]
    
class ActionCheckStatus(Action):
    def name(self) -> str:
        return "action_check_status"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # Get the user's message
        user_message = tracker.latest_message.get("text", "")
        
        # Use regular expression to find order ID in the user's message
        order_id_match = re.search(r'\b\d{4}\b', user_message)
        print(order_id_match)
        if order_id_match:
            # Extract the order ID
            order_id = order_id_match.group()
            print(order_id)
            # Retrieve order details from the Excel sheet
            order_details = self.retrieve_order_details(order_id)
            print(order_details)
            if order_details:
                # Extract the order generation time from the details
                order_generation_time = order_details.get("generation_time")
                print(order_generation_time)
                if order_generation_time:
                    # Calculate the time difference between now and order generation time
                    current_time = datetime.now()
                    time_difference = current_time - order_generation_time
                    print(time_difference)
                    # Format the time difference for display
                    formatted_time_difference = self.format_time_difference(time_difference)

                    # Check if the time difference is less than 15 minutes
                    if time_difference < timedelta(minutes=15):
                        # Calculate the remaining time until the order is ready
                        remaining_time = 15 - (time_difference.total_seconds() // 60)
                        
                        # Inform the user about the remaining time
                        dispatcher.utter_message(
                            f"The status of order ID {order_id} is: Your order will be ready in {int(remaining_time)} minutes.")
                    else:
                        # Inform the user that the order is ready for pickup
                        dispatcher.utter_message(
                            f"The status of order ID {order_id} is: Your order is ready for pickup!")
                else:
                    # Inform the user if the order generation time cannot be retrieved
                    dispatcher.utter_message(
                        "Error: Unable to retrieve order generation time.")
            else:
                # Inform the user if the order ID is not found in the Excel sheet
                dispatcher.utter_message(
                    f"Error: Order ID {order_id} not found.")
        else:
            # Inform the user if no order ID is provided in the message
            dispatcher.utter_message("Error: Order ID not provided in the message.")

        return []

    def retrieve_order_details(self, order_id: str) -> Dict[Text, Any]:
        try:
            # Load the Excel workbook
            workbook = openpyxl.load_workbook("food_order.xlsx")
            
            # Access the active sheet
            sheet = workbook.active

            # Iterate through rows to find the order ID
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if str(row[0]) == order_id:
                    # If found, return order details
                    order_details = {
                        "generation_time": row[5]  # Assuming the order generation time is in the seventh column
                    }
                    workbook.close()
                    return order_details
        except FileNotFoundError:
            pass

        # Return an empty dictionary if order details cannot be retrieved
        return {}

    def format_time_difference(self, time_difference: timedelta) -> str:
        # Extract days and seconds from the time difference
        seconds =  time_difference.seconds

        # Convert seconds to hours and minutes
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60

        # Format the time for display
        formatted_time = f"{hours} hours, {minutes} minutes"
        return formatted_time
    
class ActionShowMenu(Action):
    def name(self) -> str:
        return "action_show_menu"

    def run(self, dispatcher: CollectingDispatcher, tracker: Tracker, domain: Dict[Text, Any]) -> List[Dict[Text, Any]]:
        # Establish a connection to the database
        connection = mc.connect(
            host="localhost",
            user="Rasa_database",
            password="Kutub@123",
            database="rasa"
        )
        cursor = connection.cursor()
        
        # Get the user's message
        latest_message = tracker.latest_message
        text = latest_message.get('text', '')
        print(text)
        # Determine the category of menu requested by the user
        if "fastfood" in text.lower():
            category = "FastFood"
        elif "southindian" in text.lower():
            category = "SouthIndian"
        else:
            # If the user doesn't specify a category, show the entire menu
            category = None

        # Execute the SQL query to fetch the menu based on category if specified
        if category:
            cursor.execute("SELECT * FROM menu WHERE categories=%s", (category,))
        else:
            cursor.execute("SELECT * FROM menu")
            
        data = cursor.fetchall()

        # Close the cursor and connection
        cursor.close()
        connection.close()

        # Format the menu data as a table using Markdown
        menu_message = f"Here's our {category or 'menu'}:\n\n```\n"
        menu_message += "+----+--------------+-------+\n"
        menu_message += "| ID | Food-items   | Price |\n"
        menu_message += "+----+--------------+-------+\n"

        for item in data:
            menu_message += f"| {item[0]:2} | {item[1]:<12} | {item[2]:>5} |\n"

        menu_message += "+----+--------------+-------+\n"
        menu_message += "```"

        dispatcher.utter_message(menu_message)

        return []

#rasa run -m models --enable-api --cors "*" --debug